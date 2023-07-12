from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import customtkinter as ctk
from tkinter import ttk


class MainApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self._set_appearance_mode("dark")

        self.geometry("1000x600")

        self.open_button = ctk.CTkButton(
            self, command=self.open_excel, text="Open excel")
        self.open_button.place(relx=0.5, rely=0.8, anchor="center")
        self.excel = Excel(self)

        self.make_v_scrollbar()
        self.make_h_scrollbar()
        # self.bind("<FocusOut>", exit)

        self.mainloop()

    def open_excel(self):
        path = ctk.filedialog.askopenfile(filetypes=[("Excel", "*.xlsx")])
        
        if path is not None:
            self.excel.clear()
            wb = load_workbook(path.name)
            ws = wb.worksheets[0]
            while ws.max_column > self.excel.column_num:
                self.excel.increase_columns()
            while ws.max_row > self.excel.row_num:
                self.excel.increase_columns()
            for row in range(1, ws.max_row + 1):
                for column in range(1, ws.max_column + 1):
                    pos = f"{get_column_letter(column)}{row}"
                    value = ws[pos].value
                    self.excel.set(row, column=column, value=value)

    def v_scrollbar_command(self, *scroll):
        if float(scroll[1]) > 0.95:
            self.excel.increase_rows()
            
        self.v_scrollbar.set(*scroll)

    def make_v_scrollbar(self):
        self.v_scrollbar = ctk.CTkScrollbar(self, command=self.excel.yview)
        self.v_scrollbar.place(relx=1, y=0, relheight=0.7, anchor="ne")
        self.excel.configure(yscrollcommand=self.v_scrollbar_command)

    def h_scrollbar_command(self, *scroll):
        if float(scroll[1]) > 0.95:
            self.excel.increase_columns()
        self.h_scrollbar.set(*scroll)

    def make_h_scrollbar(self):
        self.h_scrollbar = ctk.CTkScrollbar(self, command=self.excel.xview, orientation=ctk.HORIZONTAL)
        self.h_scrollbar.place(x=0, rely=0.7, relwidth=1, anchor="nw")
        self.excel.configure(xscrollcommand=self.h_scrollbar_command)


class Excel(ttk.Treeview):
    def __init__(self, parent):
        super().__init__(parent, show="headings", padding=(0, 0, 0, 0))
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Treeview",
            background='#707070',
            fieldbackground="#707070",
            foreground="#D7D7D7",
            font=(15))
        
        style.configure(
            "Treeview.Heading",
            background='#707070',
            relief=ctk.SOLID,
            bordercolor="black")
        
        nums = tuple(range(27))
        self["columns"] = nums

        self.column(0, stretch=False, width=50)
        self.row_num = 26
        self.column_num = 26
        for col in nums[1:]:
            self.column(col, stretch=False, width=100)
            self.heading(col, text=get_column_letter(col))
            self.insert("", index="end", iid=col)
            self.set(col, 0, col)

        self.place(x=0, y=0, relwidth=1, relheight=0.7)

    def increase_rows(self):
        row = self.row_num + 1
        self.insert("", index="end", iid=row)
        self.set(row, 0, row)
        self.row_num = row

    def increase_columns(self):
        col = self.column_num + 1
        self["columns"] = (*self["columns"], (str(col)))
        for col in range(1, int(self["columns"][-1]) + 1):

            self.column(col, width=100, stretch=False)
            self.heading(col, text=get_column_letter(col))
        self.column_num = col
    

    def clear(self):
        self.delete(*self.get_children())
 
        for row in range(1, self.row_num + 1):
            self.insert("", index="end", iid=row)
            self.set(row, 0, row)

MainApp()
